from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, Http404, FileResponse
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from datetime import timezone as dt_timezone
from django.db.models import Q
from django.template.loader import render_to_string
import mimetypes
import os
import openpyxl
from openpyxl.utils import get_column_letter

from .models import Room, Activity, Submission, Notification, Announcement
from user.models import StudentProfile, TeacherProfile
from .notifications import notify_student_enrolled, notify_student_left, notify_student_submission, notify_activity_graded, notify_new_activity

from .utils.grades import calculate_grade


@login_required
def room_view(request, room_id):
    Activity.close_past_due_bulk()
    if request.user.is_authenticated:
        room = get_object_or_404(Room, id=room_id)
        
        breadcrumb_items = [
            {'text': 'Dashboard', 'url': '/dashboard/', 'icon': 'bi bi-house'},
            {'text': 'My Rooms', 'url': '/room/all/', 'icon': 'bi bi-collection-play'},
            {'text': room.name, 'url': '', 'icon': 'bi bi-door-open'}
        ]
        
        context = {
            'room': room,
            'activities': Activity.objects.filter(room=room).order_by('-created_at'),
            'announcements': Announcement.objects.filter(room=room).order_by('-created_at'),
            'breadcrumb_items': breadcrumb_items
        }

        if hasattr(request.user, 'teacher'):
            students_with_grades = []
            for student in room.students.all():
                try:
                    student_profile = StudentProfile.objects.get(user=student)
                    submissions = Submission.objects.filter(student=student_profile, activity__room=room)
                    total_score = 0
                    total_possible = 0
                    
                    for submission in submissions:
                        if submission.score is not None:
                            total_score += submission.score
                            total_possible += submission.activity.total_marks
                    
                    overall_percent, overall_grade = calculate_grade(total_score, total_possible, base_passing=room.base_passing)
                    
                    students_with_grades.append({
                        'id': student.id,
                        'student': student,
                        'grade': overall_percent,
                        'submissions_count': submissions.count(),
                        'graded_submissions': submissions.filter(score__isnull=False).count()
                    })
                except StudentProfile.DoesNotExist:
                    students_with_grades.append({
                        'id': student.id,
                        'student': student,
                        'grade': 0,
                        'submissions_count': 0,
                        'graded_submissions': 0
                    })
            
            students_content = render_to_string('room/components/student_list.html', {
                'students_with_grades': students_with_grades,
                'user_type': 'teacher'
            })
            
            activities_content = render_to_string('room/components/activities_list.html', {
                'activities': context['activities'],
                'user_type': 'teacher'
            })
            
            announcements_content = render_to_string('room/components/announcements_list.html', {
                'announcements': context['announcements'],
                'user_type': 'teacher'
            })
            
            context.update({
                'students_with_grades': students_with_grades,
                'students_content': students_content,
                'activities_content': activities_content,
                'announcements_content': announcements_content
            })
            return render(request, 'room/teacher.html', context)
        
        elif hasattr(request.user, 'student'):
            try:
                student_profile = StudentProfile.objects.get(user=request.user)
            except StudentProfile.DoesNotExist:
                student_profile = None

            activities = context['activities']
            submissions = []
            if student_profile:
                submissions = list(Submission.objects.filter(activity__in=activities, student=student_profile))
            submissions_by_activity = {s.activity_id: s for s in submissions}

            pending_count = 0
            scored_sum = 0
            possible_sum = 0
            for activity in activities:
                submission = submissions_by_activity.get(activity.id)
                if submission is None or submission.score is None:
                    pending_count += 1
                if submission is not None and submission.score is not None:
                    scored_sum += int(submission.score)
                    possible_sum += int(activity.total_marks or 0)

            overall_percent, overall_grade = calculate_grade(scored_sum, possible_sum, base_passing=room.base_passing)

            students_with_grades = [{
                'student': request.user,
                'grade': overall_percent,
                'submissions_count': len(submissions),
                'graded_submissions': len([s for s in submissions if s.score is not None])
            }]

            activities_content = render_to_string('room/components/activities_list.html', {
                'activities': context['activities'],
                'user_type': 'student',
                'submissions_by_activity': submissions_by_activity
            })
            
            announcements_content = render_to_string('room/components/announcements_list.html', {
                'announcements': context['announcements'],
                'user_type': 'student'
            })

            context.update({
                'student_pending_activities': pending_count,
                'student_overall_percent': overall_percent,
                'students_with_grades': students_with_grades,
                'activities_content': activities_content,
                'announcements_content': announcements_content,
                'submissions_by_activity': submissions_by_activity,
            })

            return render(request, 'room/student.html', context)
        
    return redirect('home')

@login_required
def view_all_room(request):
    if request.user.is_authenticated:
        room = []
        q = (request.GET.get('q') or '').strip()
        if hasattr(request.user, 'teacher'):
            room = Room.objects.filter(teacher__user=request.user)
        elif hasattr(request.user, 'student'):
            room = Room.objects.filter(students=request.user)
        else:
            return redirect('role_selection')
        
        if q:
            from django.db.models import Q
            room = room.filter(Q(name__icontains=q) | Q(room_code__icontains=q))

        breadcrumb_items = [
            {'text': 'Dashboard', 'url': '/dashboard/', 'icon': 'bi bi-house'},
            {'text': 'My Rooms', 'url': '', 'icon': 'bi bi-collection-play'}
        ]

        context = {
            'rooms': room,
            'breadcrumb_items': breadcrumb_items,
            'q': q,
        }

        return render(request, 'rooms.html', context)

    return redirect('home')

@login_required
def enroll_student(request):
    if request.method == 'POST':
        room_code = request.POST.get('code')
        room = Room.objects.filter(room_code=room_code.strip()).first()

        if room and hasattr(request.user, 'student'):
            room.students.add(request.user)
            room.save()
            notify_student_enrolled(room, request.user)
            messages.success(request, f"You have successfully enrolled in {room.name}.")
            return redirect('room', room_id=room.id)

        messages.error(request, "Invalid room code or you are not a student.")

    return redirect('all_room')

@login_required
def activity_view(request, activity_id):
    activity = get_object_or_404(Activity, id=activity_id)
    room = activity.room
    breadcrumb_items = [
        {'text': 'Dashboard', 'url': '/dashboard/', 'icon': 'bi bi-house'},
        {'text': 'My Rooms', 'url': '/room/all/', 'icon': 'bi bi-collection-play'},
        {'text': room.name, 'url': f'/room/{room.id}/', 'icon': 'bi bi-door-open'},
        {'text': 'Activity', 'url': '', 'icon': 'bi bi-journal-text'},
    ]

    if activity.due_date and activity.due_date < timezone.now() and activity.status != 'closed':
        activity.status = 'closed'
        activity.save(update_fields=['status'])

    context = {
        'breadcrumb_items': breadcrumb_items,
        'submission': None,
        'activity': activity,
    }

    if hasattr(request.user, 'student'):
        try:
            student = StudentProfile.objects.get(user=request.user)
            submission = Submission.objects.get(activity=activity, student=student)
            context['submission'] = submission
        except Submission.DoesNotExist:
            context['submission'] = None
        return render(request, 'activity/student.html', context)
    elif hasattr(request.user, 'teacher'):
        submissions = Submission.objects.filter(activity=activity)
        context['submissions'] = submissions
        return render(request, 'activity/teacher.html', context)
    return redirect('all_room')

@login_required
def announcement_view(request, announcement_id):
    announcement = get_object_or_404(Announcement, id=announcement_id)
    room = announcement.room
    breadcrumb_items = [
        {'text': 'Dashboard', 'url': '/dashboard/', 'icon': 'bi bi-house'},
        {'text': 'My Rooms', 'url': '/room/all/', 'icon': 'bi bi-collection-play'},
        {'text': room.name, 'url': f'/room/{room.id}/', 'icon': 'bi bi-door-open'},
        {'text': 'Announcement', 'url': '', 'icon': 'bi bi-megaphone'}
    ]
    
    context = {
        'announcement': announcement,
        'breadcrumb_items': breadcrumb_items
    }

    return render(request, 'announcement.html', context)

@login_required
def create_room(request):
    if request.method == 'POST':
        
        if request.user.is_authenticated and hasattr(request.user, 'teacher'):
            name = request.POST.get('name')
            description = request.POST.get('description')
            teacher = TeacherProfile.objects.get(user=request.user)

            room = Room.objects.create(
                name=name,
                description=description,
                teacher=teacher
            )
            room.save()
            return redirect('room', room_id=room.id)

    return redirect('all_room')

@login_required
def delete_room(request, room_id):
    if request.user.is_authenticated and hasattr(request.user, 'teacher'):
        room = Room.objects.get(id=room_id)
        if room.teacher.user == request.user:
            room.delete()
    return redirect('all_room')

@login_required
def create_activity(request):
    if request.method == 'POST' and request.user.is_authenticated and hasattr(request.user, 'teacher'):
        title = request.POST.get('title')
        description = request.POST.get('description')
        due_date_input = request.POST.get('deadline')
        room_id = request.POST.get('room_id')
        total_marks = request.POST.get('total_points')
        resource_file = request.FILES.get('resource')

        room = Room.objects.get(id=room_id)

        due_date_value = None
        if due_date_input:
            parsed = parse_datetime(due_date_input)
            if parsed is None:
                try:
                    from datetime import datetime
                    parsed = datetime.fromisoformat(due_date_input)
                except Exception:
                    parsed = None
            if parsed is not None:
                if timezone.is_naive(parsed):
                    parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
                due_date_value = parsed.astimezone(dt_timezone.utc)

        activity = Activity.objects.create(
            title=title,
            description=description,
            due_date=due_date_value,
            room=room,
            total_marks=total_marks,
            resource_file=resource_file if resource_file else None
        )
        
        notify_new_activity(activity)

        return redirect('room', room_id=room.id)

    return redirect('all_room')

@login_required
def create_announcement(request):
    if request.method == 'POST':

        if request.user.is_authenticated and hasattr(request.user, 'teacher'):
            title = request.POST.get('title')
            content = request.POST.get('content')
            room_id = request.POST.get('room_id')

            room = Room.objects.get(id=room_id)

            announcement = Announcement.objects.create(
                title=title,
                content=content,
                room=room
            )
            announcement.save()
            
            from .notifications import create_student_notifications
            create_student_notifications(
                students=room.students.all(),
                notification_type='new_announcement',
                title=f'New Announcement: {title}',
                message=f'A new announcement has been posted in {room.name}',
                room=room
            )
            
            return redirect('room', room_id=room.id)

    return redirect('all_room')

@login_required
def submit_activity(request):
    if request.method == 'POST':
        if request.user.is_authenticated and hasattr(request.user, 'student'):
            activity_id = request.POST.get('activity_id')
            submission_file = request.FILES.get('submission_file')

            if activity_id and submission_file:
                activity = get_object_or_404(Activity, id=activity_id)
                student = get_object_or_404(StudentProfile, user=request.user)

                if (activity.due_date and activity.due_date <= timezone.now()) or activity.status == 'closed':
                    return redirect('activity_view', activity_id=activity_id)

                submission, created = Submission.objects.get_or_create(
                    activity=activity,
                    student=student,
                    defaults={'submission_file': submission_file}
                )

                if not created:
                    if submission.submission_file:
                        submission.submission_file.delete(save=False)
                    submission.submission_file = submission_file
                    submission.save()

                if activity.status != 'closed':
                    submission.status = 'submitted'
                    submission.save()
                    
                    notify_student_submission(submission)
                
                return redirect('activity_view', activity_id=activity_id)

    return redirect('all_room')

@login_required
def grade_submission(request):
    if request.method == 'POST':
        if request.user.is_authenticated and hasattr(request.user, 'teacher'):
            submission_id = request.POST.get('submission_id')
            score = request.POST.get('score')
            feedback = request.POST.get('feedback')

            submission = get_object_or_404(Submission, id=submission_id)
            activity = submission.activity
            room = activity.room

            if room.teacher.user == request.user:
                try:
                    score_value = int(score) if score else 0
                    max_score = activity.total_marks
                    
                    if score_value < 0:
                        messages.error(request, "Score cannot be negative.")
                        return redirect('activity_view', activity_id=activity.id)
                    if score_value > max_score:
                        messages.error(request, f"Score cannot exceed the maximum score of {max_score} points.")
                        return redirect('activity_view', activity_id=activity.id)
                    
                    submission.score = score_value
                    submission.feedback = feedback
                    submission.status = 'graded'
                    submission.save()
                    
                    notify_activity_graded(submission)
                    
                    messages.success(request, f"Successfully graded submission with {score_value}/{max_score} points.")
                    return redirect('activity_view', activity_id=activity.id)
                    
                except ValueError:
                    messages.error(request, "Please enter a valid numeric score.")
                    return redirect('activity_view', activity_id=activity.id)
    return redirect('all_room')

@login_required
def leave_room(request, room_id):
    if request.user.is_authenticated and hasattr(request.user, 'student'):
        room = Room.objects.get(id=room_id)
        room.students.remove(request.user)
        room.save()
        notify_student_left(room, request.user)
        return redirect('all_room')
    return redirect('home')

@login_required
def unenroll_student(request):
    if request.method == 'POST':
        if hasattr(request.user, 'teacher'):
            room_id = request.POST.get('room_id')
            student_id = request.POST.get('student_id')

            room = Room.objects.get(id=room_id)
            student = User.objects.get(id=student_id)

            if student in room.students.all():
                room.students.remove(student)
                room.save()
                notify_student_left(room, student)
                messages.success(request, f"Student {student.username} has been unenrolled from the room.")
                return redirect('room', room_id=room.id)
            else:
                messages.error(request, f"Student {student.username} is not enrolled in this room.")
                
    return redirect('all_room')

@login_required
def export_grades(request, room_id):
    if not hasattr(request.user, 'teacher'):
        messages.error(request, "Only teachers can export grades.")
        return redirect('all_room')

    try:
        room = Room.objects.get(id=room_id)
    except Room.DoesNotExist:
        raise Http404("Room not found")

    if room.teacher.user != request.user:
        messages.error(request, "You had no access to export grades from this room!")
        return redirect('all_room')

    work_book = openpyxl.Workbook()
    work_sheet = work_book.active
    work_sheet.title = room.name.title()

    headers = [
        'Student ID',
        'Last Name',
        'First Name',
        'Middle Name',
        'Course',
        'Year Level',
        'Subject',
    ]
    for activity in room.activities.all():
        headers.append(activity.title)
    headers += ['Total Score', 'Max Score', 'Average %']

    work_sheet.append(headers)

    for user in sorted(room.students.all(), key=lambda user: (user.last_name.lower(), user.first_name.lower())):
        student = user.student
        row = [
            student.student_id,
            user.last_name,
            user.first_name,
            student.middle_name if student.middle_name else '-',
            student.course,
            student.year_level.split(' ')[0],
            room.name,
        ]

        total_score = 0
        max_score = 0

        for activity in room.activities.all():
            submission = Submission.objects.filter(activity=activity, student=student).first()
            score = submission.score if submission else 0
            row.append(score if score != 0 else "-")

            total_score += score
            max_score += activity.total_marks

        row.append(total_score)
        row.append(max_score)
        average_percent = (total_score / max_score * 100) if max_score > 0 else 0
        row.append(round(average_percent, 2))

        work_sheet.append(row)

    for col in work_sheet.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = (max_length + 2)
        work_sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{room.name.upper()} - grades.xlsx"'
    work_book.save(response)

    return response


@login_required
def serve_activity_resource(request, activity_id):
    activity = get_object_or_404(Activity, id=activity_id)
    
    if hasattr(request.user, 'teacher'):
        if activity.room.teacher.user != request.user:
            raise Http404("File not found")
    elif hasattr(request.user, 'student'):
        if request.user not in activity.room.students.all():
            raise Http404("File not found")
    else:
        raise Http404("File not found")
    
    if not activity.resource_file:
        raise Http404("No resource file available")
    
    try:
        file_path = activity.resource_file.path
        
        content_type, _ = mimetypes.guess_type(file_path)
        if content_type is None:
            content_type = 'application/octet-stream'
        
        response = FileResponse(
            open(file_path, 'rb'),
            content_type=content_type
        )
        
        filename = activity.get_resource_filename()
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        
        return response
    except FileNotFoundError:
        raise Http404("File not found")


@login_required
def serve_submission_file(request, submission_id):
    submission = get_object_or_404(Submission, id=submission_id)
    
    if hasattr(request.user, 'teacher'):
        if submission.activity.room.teacher.user != request.user:
            raise Http404("File not found")
    elif hasattr(request.user, 'student'):
        if submission.student.user != request.user:
            raise Http404("File not found")
    else:
        raise Http404("File not found")
    
    if not submission.submission_file:
        raise Http404("No submission file available")
    
    try:
        file_path = submission.submission_file.path
        
        content_type, _ = mimetypes.guess_type(file_path)
        if content_type is None:
            content_type = 'application/octet-stream'
        
        response = FileResponse(
            open(file_path, 'rb'),
            content_type=content_type
        )
        
        filename = submission.get_submission_filename()
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        
        return response
    except FileNotFoundError:
        raise Http404("File not found")


@login_required
def notifications_view(request):
    notifications = Notification.objects.filter(recipient=request.user)
    unread_count = notifications.filter(is_read=False).count()
    
    context = {
        'notifications': notifications,
        'unread_count': unread_count,
        'breadcrumb_items': [
            {'text': 'Dashboard', 'url': '/dashboard/', 'icon': 'bi bi-house'},
            {'text': 'Notifications', 'url': '', 'icon': 'bi bi-bell'}
        ]
    }
    return render(request, 'notifications.html', context)


@login_required
def mark_notification_read(request, notification_id):
    try:
        notification = Notification.objects.get(id=notification_id, recipient=request.user)
        notification.is_read = True
        notification.save()
        
        target_url = notification.get_url()
        if target_url and target_url != '/room/all/':
            return redirect(target_url)
    except Notification.DoesNotExist:
        pass
    
    return redirect('notifications')


@login_required
def mark_all_notifications_read(request):
    Notification.objects.filter(recipient=request.user, is_read=False).update(is_read=True)
    messages.success(request, "All notifications marked as read.")
    return redirect('notifications')


@login_required
def clear_all_notifications(request):
    if request.method == 'POST':
        count = Notification.objects.filter(recipient=request.user).count()
        Notification.objects.filter(recipient=request.user).delete()
        messages.success(request, f"All {count} notifications have been cleared.")
    return redirect('notifications')
